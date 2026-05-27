#!/usr/bin/env python3

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.stderr.write("Creating the editor workbook requires openpyxl.\n")
    sys.exit(2)


SEED_HEADERS = ["am_email", "am_name", "account_name", "domain", "status", "notes"]
STATUSES = ["domain_pending", "ready_for_intake", "identity_review", "hold"]


def main():
    seed_path = Path(sys.argv[1] if len(sys.argv) > 1 else "templates/am-account-seed-list.csv")
    roster_path = Path(sys.argv[2] if len(sys.argv) > 2 else "templates/am-roster.csv")
    output_path = Path(sys.argv[3] if len(sys.argv) > 3 else "/private/tmp/myra-am-organization-editor.xlsx")

    seeds = read_dicts(seed_path)
    roster = read_dicts(roster_path)

    workbook = Workbook()
    workbook.remove(workbook.active)

    build_instructions(workbook)
    build_editor(workbook, seeds, roster)
    build_roster(workbook, roster)
    build_summary(workbook, roster)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Created {output_path}")


def read_dicts(path):
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def build_instructions(workbook):
    sheet = workbook.create_sheet("Start Here")
    sheet["A1"] = "myRA AM Organization Assignment Editor"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F2937")
    sheet.merge_cells("A1:F1")

    rows = [
        ("Purpose", "Edit which organizations/accounts belong to each AM before regenerating packages or writing to Day AI."),
        ("Edit sheet", "Use the 'Edit Assignments' tab. Change AM email/name, account name, domain, status, and notes."),
        ("AM names", "Use the dropdown values from the 'AM Roster' tab. Keep email/name paired correctly."),
        ("Domain", "Leave blank only when status is domain_pending or identity_review."),
        ("Status", "Use domain_pending, ready_for_intake, identity_review, or hold."),
        ("Export", "After editing, run: npm run org-editor:import -- /path/to/workbook.xlsx"),
        ("Validate", "Then run: npm run validate:account-seeds"),
        ("Package", "Then run: npm run am:tour-packages"),
        ("Day AI", "Use Day AI writes only after reviewing the generated provisioning payload/receipts."),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(index, 1, label).font = Font(bold=True)
        sheet.cell(index, 2, value)

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows(min_row=1, max_row=12, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_editor(workbook, seeds, roster):
    sheet = workbook.create_sheet("Edit Assignments")
    sheet.append(SEED_HEADERS)

    for row in seeds:
        sheet.append([row.get(header, "") for header in SEED_HEADERS])

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max(sheet.max_row, 2)}"
    widths = [32, 24, 42, 28, 20, 58]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[column_letter(index)].width = width

    roster_count = max(len(roster), 1)
    email_validation = DataValidation(type="list", formula1=f"'AM Roster'!$A$2:$A${roster_count + 1}", allow_blank=False)
    status_validation = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=False)
    sheet.add_data_validation(email_validation)
    sheet.add_data_validation(status_validation)
    email_validation.add(f"A2:A{max(sheet.max_row + 100, 300)}")
    status_validation.add(f"E2:E{max(sheet.max_row + 100, 300)}")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_roster(workbook, roster):
    sheet = workbook.create_sheet("AM Roster")
    headers = ["am_email", "am_name"]
    sheet.append(headers)
    for row in roster:
        sheet.append([row.get("am_email", ""), row.get("am_name", "")])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCFCE7")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{max(sheet.max_row, 2)}"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 28


def build_summary(workbook, roster):
    sheet = workbook.create_sheet("Summary")
    sheet["A1"] = "Assignment Counts"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet.append(["AM Email", "AM Name", "Assigned Accounts"])
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FDE68A")

    for index, row in enumerate(roster, start=3):
        email = row.get("am_email", "")
        sheet.cell(index, 1, email)
        sheet.cell(index, 2, row.get("am_name", ""))
        sheet.cell(index, 3, f'=COUNTIF(\'Edit Assignments\'!$A:$A,A{index})')

    total_row = len(roster) + 4
    sheet.cell(total_row, 2, "Total").font = Font(bold=True)
    sheet.cell(total_row, 3, f"=SUM(C3:C{len(roster) + 2})").font = Font(bold=True)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 20


def column_letter(index):
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


if __name__ == "__main__":
    main()
