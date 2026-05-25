#!/usr/bin/env python3

import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "Reading .xlsx files requires Python package openpyxl. "
        "Install openpyxl or export the sheet as CSV.\n"
    )
    sys.exit(2)

if len(sys.argv) != 2:
    sys.stderr.write("Usage: scripts/read-xlsx.py <file.xlsx>\n")
    sys.exit(2)

workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
worksheet = workbook[workbook.sheetnames[0]]

rows = []
for row in worksheet.iter_rows(values_only=True):
    rows.append(["" if value is None else str(value) for value in row])

print(json.dumps(rows))

