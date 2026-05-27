#!/usr/bin/env python3

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Importing the editor workbook requires openpyxl.\n")
    sys.exit(2)


HEADERS = ["am_email", "am_name", "account_name", "domain", "status", "notes"]


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("Usage: scripts/import-organization-editor.py <editor.xlsx> [output.csv]\n")
        sys.exit(2)

    workbook_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2] if len(sys.argv) > 2 else "templates/am-account-seed-list.csv")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Edit Assignments" not in workbook.sheetnames:
        sys.stderr.write("Workbook must contain an 'Edit Assignments' sheet.\n")
        sys.exit(1)

    sheet = workbook["Edit Assignments"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        sys.stderr.write("'Edit Assignments' sheet is empty.\n")
        sys.exit(1)

    headers = [clean(value) for value in rows[0]]
    missing = [header for header in HEADERS if header not in headers]
    if missing:
        sys.stderr.write(f"Missing required columns: {', '.join(missing)}\n")
        sys.exit(1)

    indexes = [headers.index(header) for header in HEADERS]
    output_rows = []
    for raw_row in rows[1:]:
        values = [clean(raw_row[index]) if index < len(raw_row) else "" for index in indexes]
        record = dict(zip(HEADERS, values))
        if not any(record.values()):
            continue
        if not record["account_name"]:
            continue
        output_rows.append(record)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(output_rows)

    print(f"Wrote {len(output_rows)} assignment row(s) to {output_path}")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    main()
