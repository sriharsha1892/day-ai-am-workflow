#!/usr/bin/env python3

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "Writing .xlsx files requires Python package openpyxl. "
        "Run this script from a Codex workspace with bundled spreadsheet dependencies.\n"
    )
    sys.exit(2)


if len(sys.argv) != 3:
    sys.stderr.write("Usage: scripts/write-am-workbook.py <account-packet.json> <MY_ACCOUNTS.xlsx>\n")
    sys.exit(2)

packet_path = Path(sys.argv[1])
output_path = Path(sys.argv[2])
packet = json.loads(packet_path.read_text())
accounts = packet["accounts"]
active_contacts = packet.get("activeContacts", [])
am = packet["am"]
summary = packet["summary"]

wb = Workbook()
default = wb.active
wb.remove(default)

fills = {
    "header": PatternFill("solid", fgColor="1F2937"),
    "blue": PatternFill("solid", fgColor="DBEAFE"),
    "green": PatternFill("solid", fgColor="DCFCE7"),
    "yellow": PatternFill("solid", fgColor="FEF3C7"),
    "orange": PatternFill("solid", fgColor="FFEDD5"),
    "gray": PatternFill("solid", fgColor="E5E7EB"),
    "white": PatternFill("solid", fgColor="FFFFFF"),
}
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_cell(sheet, cell, value):
    sheet[cell] = value
    sheet[cell].font = Font(size=18, bold=True, color="111827")


def section_header(sheet, row, values):
    for col, value in enumerate(values, 1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.fill = fills["header"]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def style_table(sheet, min_row, max_row, min_col, max_col):
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet, widths):
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


start = wb.create_sheet("Start Here")
title_cell(start, "A1", "myRA AM Tour")
start["A3"] = "AM"
start["B3"] = am["name"]
start["A4"] = "Email"
start["B4"] = am["email"]
start["A6"] = "Open Codex in this folder and say:"
start["A7"] = packet["tour"]["startPrompt"]
start["A7"].font = Font(size=14, bold=True, color="1D4ED8")
start["D3"] = "Tour Modes"
start["D3"].font = Font(bold=True)
mode_rows = [
    ("Beginner", packet["tour"].get("beginnerPrompt", "Start my myRA AM tour in beginner mode.")),
    ("Standard", packet["tour"].get("standardPrompt", packet["tour"]["startPrompt"])),
    ("Power", packet["tour"].get("powerPrompt", "Start my myRA AM tour in power mode.")),
]
for row_index, (mode, prompt) in enumerate(mode_rows, 4):
    start.cell(row=row_index, column=4, value=mode).font = Font(bold=True)
    start.cell(row=row_index, column=5, value=prompt)
start["A9"] = "What Codex will do"
start["A9"].font = Font(bold=True)
actions = [
    "Check Day AI MCP access.",
    "Load account-packet.json for speed.",
    "Load the myRA context pack for every recommendation.",
    "Show your priority queue.",
    "Recommend the next account.",
    "Run smart Organization matching before intake writes.",
    "Request Freshsales/Apollo/Clearout through centralized connectors when needed.",
    "Pause before every Day AI write.",
    "Show a Day AI receipt after every write.",
]
for offset, action in enumerate(actions, 10):
    start.cell(row=offset, column=1, value=f"- {action}")
start["D9"] = "What Codex will not do"
start["D9"].font = Font(bold=True)
guardrails = [
    "It will not send emails.",
    "It will not write to Freshsales.",
    "It will not create contacts without approval.",
    "It will not create duplicate Organizations when a match is ambiguous.",
    "It will not use provider keys from this package.",
]
for offset, action in enumerate(guardrails, 10):
    start.cell(row=offset, column=4, value=f"- {action}")
set_widths(start, [24, 34, 4, 28, 42])
start.freeze_panes = "A9"

today = wb.create_sheet("Today")
title_cell(today, "A1", "Today's Tour")
recommended = next((account for account in accounts if account["status"] == "ready_for_intake"), accounts[0] if accounts else None)
recommended_reason = recommended.get("recommendedReason", "") if recommended else ""
today_rows = [
    ("Recommended account", recommended["accountName"] if recommended else "None ready"),
    ("Domain", recommended.get("domain", "") if recommended else ""),
    ("Priority", recommended.get("priority", "") if recommended else ""),
    ("Why this account", recommended_reason or "Ready accounts are prioritized by status, priority, then packet order."),
    ("Mode to use", f"Default: {packet['tour'].get('defaultMode', 'standard')} | Beginner available for step-by-step guidance"),
    ("Start prompt", packet["tour"]["startPrompt"]),
    ("Step 1: Account safety", recommended.get("orgResolutionCommand", "") if recommended else ""),
    ("Step 2: Intake only if safe", recommended.get("intakeCommand", "") if recommended else ""),
    ("Step 3: Research", f'/research-account domain="{recommended.get("domain", "")}"' if recommended else ""),
    ("Step 4: Contacts", f'/map-contacts domain="{recommended.get("domain", "")}"' if recommended else ""),
    ("Step 5: Health", f'/account-health domain="{recommended.get("domain", "")}"' if recommended else ""),
    ("Stop condition", "If org match is Red/ambiguous, do not create an Organization. Create review context only."),
    ("Day AI writes", "Org/opportunity/context only after approval; People/drafts/actions require later checkpoints."),
]
for row_index, (label, value) in enumerate(today_rows, 3):
    today.cell(row=row_index, column=1, value=label).font = Font(bold=True)
    today.cell(row=row_index, column=2, value=value)
style_table(today, 3, len(today_rows) + 2, 1, 2)
set_widths(today, [24, 95])

commands = wb.create_sheet("Command Cards")
title_cell(commands, "A1", "Command Cards")
section_header(commands, 3, ["I want to...", "Say this to Codex", "What happens", "Day AI write?"])
command_rows = [
    ("Start guided tour", packet["tour"]["startPrompt"], "Shows queue, recommends account, starts five-station path.", "No write until approval"),
    ("Use step-by-step mode", packet["tour"].get("beginnerPrompt", "Start my myRA AM tour in beginner mode."), "Codex explains each checkpoint and asks one decision at a time.", "No write until approval"),
    ("Check duplicate account", "Run smart org match for this account before intake.", "Runs org-resolution and shows Green/Yellow/Red receipt.", "Review/link/create only after gate"),
    ("Research account", "Research this account.", "Builds account brief, myRA use cases, signals, buyer hypothesis.", "Context/page after approval"),
    ("Find ICP", "Find ICP for this account.", "Maps likely personas and role buckets for myRA value.", "No People write"),
    ("Find leads", "Find leads for this account.", "Combines imported contacts, Apollo, Freshsales, and Day AI candidates.", "No People write"),
    ("Build cadence", "Build my cadence.", "Creates branching outreach plan for selected contacts.", "Actions/drafts after approval"),
    ("Draft first email", "Write first email.", "Creates a reviewable email draft grounded in account signal and myRA use case.", "Draft after approval"),
    ("Check saved work", "Show what has been saved to Day AI.", "Shows Day AI receipts, pending sync, next action.", "Optional snapshot"),
    ("Recover from crash", "Retry pending Day AI sync for this account using the same idempotency key.", "Retries the same intended write without duplicating orgs.", "Retry only"),
]
for row_index, row in enumerate(command_rows, 4):
    for col, value in enumerate(row, 1):
        commands.cell(row=row_index, column=col, value=value)
style_table(commands, 3, len(command_rows) + 3, 1, 4)
set_widths(commands, [28, 58, 58, 26])

queue = wb.create_sheet("My Queue")
queue_headers = ["Priority", "Status", "Account Name", "Domain", "Confidence", "Next Action", "Notes", "Source"]
section_header(queue, 1, queue_headers)
for index, account in enumerate(accounts, 2):
    values = [
        account.get("priority", ""),
        account.get("status", ""),
        account.get("accountName", ""),
        account.get("domain", ""),
        account.get("domainConfidence", ""),
        account.get("nextAction", ""),
        account.get("notes", ""),
        account.get("domainSourceUrl", ""),
    ]
    for col, value in enumerate(values, 1):
        queue.cell(row=index, column=col, value=value)
    fill = {
        "ready_for_intake": fills["green"],
        "domain_pending": fills["yellow"],
        "identity_review": fills["orange"],
        "hold": fills["gray"],
    }.get(account.get("status"), fills["white"])
    for col in range(1, len(queue_headers) + 1):
        queue.cell(row=index, column=col).fill = fill
style_table(queue, 1, len(accounts) + 1, 1, len(queue_headers))
queue.auto_filter.ref = f"A1:H{len(accounts) + 1}"
queue.freeze_panes = "A2"
set_widths(queue, [12, 18, 34, 28, 16, 32, 48, 52])

checklist = wb.create_sheet("Tour Checklist")
title_cell(checklist, "A1", "Tour Checklist")
check_items = [
    "Day AI connected",
    "Account selected",
    "Org match checked",
    "Account intake created",
    "Research saved",
    "Contacts mapped",
    "Contacts approved",
    "Cadence created",
    "Draft created",
    "Next task created",
    "Account health reviewed",
]
section_header(checklist, 3, ["Done", "Checkpoint", "Notes"])
for index, item in enumerate(check_items, 4):
    checklist.cell(row=index, column=1, value="")
    checklist.cell(row=index, column=2, value=item)
    checklist.cell(row=index, column=3, value="")
style_table(checklist, 3, len(check_items) + 3, 1, 3)
set_widths(checklist, [12, 34, 70])

trust = wb.create_sheet("Trust Panel")
title_cell(trust, "A1", "Trust Panel Template")
section_header(trust, 3, ["Checkpoint", "Sources Used", "Confident About", "Needs AM Judgment", "Did Not Do", "Next Safest Action"])
trust_rows = [
    ("Org resolution", "Day AI, Freshsales, Apollo/org evidence, official domain", "Whether to link, ask, block, or create", "Parent/subsidiary scope", "Did not create duplicate org", "Proceed to intake only if Green or resolved Yellow"),
    ("Research", "Public sources, Day AI, Freshsales context", "myRA-fit use cases and signals", "Which use case AM wants to lead with", "Did not invent unsupported facts", "Confirm use-case thesis"),
    ("Contacts", "Imported contacts, Apollo, Freshsales, Day AI", "Recommended/Maybe/Hold candidates", "Which contacts are useful", "Did not create People", "Approve selected contacts"),
    ("Cadence", "Account context, persona pack, selected contacts", "Branching plan and next actions", "Channel/tone preference", "Did not send messages", "Approve tasks/drafts"),
    ("Draft", "Research, persona, prior touches", "Personalized draft rationale", "Final wording and sender mailbox", "Did not send email", "Review draft"),
    ("Health", "Day AI ledger/actions/context", "Stage, blockers, next action", "Whether to save snapshot", "Did not count unapproved Freshsales history", "Take next best action"),
]
for row_index, row in enumerate(trust_rows, 4):
    for col, value in enumerate(row, 1):
        trust.cell(row=row_index, column=col, value=value)
style_table(trust, 3, len(trust_rows) + 3, 1, 6)
set_widths(trust, [22, 34, 34, 34, 34, 34])

contact_cards = wb.create_sheet("Contact Cards")
title_cell(contact_cards, "A1", "Contact Review Cards")
section_header(contact_cards, 3, ["Tier", "Meaning", "AM Action", "Typical Evidence"])
card_rows = [
    ("Recommended", "Strong role fit and account evidence.", "Approve, enrich, verify, or draft.", "Title/persona match, company domain, Freshsales/Apollo/Day AI evidence"),
    ("Maybe", "Potentially useful but missing evidence.", "Keep for enrichment or ask Codex to research more.", "Weak title match, missing email, old CRM context"),
    ("Hold", "Weak fit, duplicate risk, bad email, or ambiguous company.", "Skip or send to admin review.", "Ambiguous org, invalid email, duplicate contact, low role fit"),
]
for row_index, row in enumerate(card_rows, 4):
    for col, value in enumerate(row, 1):
        contact_cards.cell(row=row_index, column=col, value=value)
style_table(contact_cards, 3, len(card_rows) + 3, 1, 4)
set_widths(contact_cards, [18, 42, 42, 58])

handoff = wb.create_sheet("Day AI Handoff")
title_cell(handoff, "A1", "Codex -> Day AI Handoff")
section_header(handoff, 3, ["Checkpoint", "Receipt Level", "Day AI Write", "Approval Needed", "Receipt Should Show"])
handoff_rows = [
    ("Org resolution", "Green/Yellow/Red", "Existing Organization link/update or review context", "Only for parent/subsidiary or ambiguity", "Match decision, confidence, candidates, evidence"),
    ("Account intake", "Green/Red", "Organization, opportunity/account motion, account context", "Yes", "Object type, name, lifecycle, link or ID"),
    ("Research", "Green/Yellow", "Account plan/context page", "Checkpoint approval", "Context/page name and next step"),
    ("Contacts", "Yellow", "Canonical People only after AM-selected approval", "Yes", "Person names, source, evidence, skipped duplicates"),
    ("Cadence", "Yellow", "Actions and email drafts", "Yes", "Task count, draft count, due dates"),
    ("Log touch", "Green/Yellow", "Ledger/context note and next action", "Yes", "Channel, outcome, next task"),
    ("Account health", "Green/Red", "Optional health snapshot/context", "Optional", "Stage, blockers, pending sync, next action"),
]
for row_index, row in enumerate(handoff_rows, 4):
    for col, value in enumerate(row, 1):
        handoff.cell(row=row_index, column=col, value=value)
style_table(handoff, 3, len(handoff_rows) + 3, 1, 5)
set_widths(handoff, [22, 18, 48, 22, 54])

help_sheet = wb.create_sheet("Help")
title_cell(help_sheet, "A1", "Help Prompts")
section_header(help_sheet, 3, ["If this happens", "Ask Codex"])
help_rows = [
    ("Day AI is not connected", "Fix my Day AI connection."),
    ("You stopped midway", "Resume my myRA AM tour."),
    ("You want proof of writes", "Show what has been saved to Day AI."),
    ("No account can start", "Show accounts needing domains."),
    ("You suspect a duplicate account", "Run smart org match for this account before intake."),
    ("You want target contacts", "Find leads for this account."),
    ("You want ICP/personas", "Find ICP for this account."),
    ("Day AI write failed", "Retry pending Day AI sync for this account using the same idempotency key."),
    ("You want to restart", "Restart from my recommended account."),
]
for row_index, row in enumerate(help_rows, 4):
    for col, value in enumerate(row, 1):
        help_sheet.cell(row=row_index, column=col, value=value)
style_table(help_sheet, 3, len(help_rows) + 3, 1, 2)
set_widths(help_sheet, [34, 64])

admin = wb.create_sheet("Admin Readiness")
title_cell(admin, "A1", "Admin Readiness")
section_header(admin, 3, ["Check", "Status", "Notes"])
admin_checks = packet.get("adminReadiness", {}).get("checks", [])
if not admin_checks:
    admin_checks = [
        {"check": "Package generated", "status": "not_started", "notes": ""},
        {"check": "Day AI connected", "status": "not_started", "notes": ""},
        {"check": "First account started", "status": "not_started", "notes": ""},
        {"check": "Org match passed", "status": "not_started", "notes": ""},
        {"check": "Day AI write confirmed", "status": "not_started", "notes": ""},
        {"check": "Blockers captured", "status": "not_started", "notes": ""},
    ]
for row_index, item in enumerate(admin_checks, 4):
    admin.cell(row=row_index, column=1, value=item.get("check", ""))
    admin.cell(row=row_index, column=2, value=item.get("status", "not_started"))
    admin.cell(row=row_index, column=3, value=item.get("notes", ""))
style_table(admin, 3, len(admin_checks) + 3, 1, 3)
set_widths(admin, [34, 20, 72])

contacts_sheet = wb.create_sheet("Active Contacts")
title_cell(contacts_sheet, "A1", "Imported Active Contacts")
contact_headers = [
    "Account",
    "Domain",
    "Contact",
    "Email",
    "Title",
    "Role Bucket",
    "Source",
    "Relationship",
    "Last Touch",
    "Next Step",
    "Selected",
    "Notes",
]
section_header(contacts_sheet, 3, contact_headers)
if active_contacts:
    for row_index, contact in enumerate(active_contacts, 4):
        values = [
            contact.get("accountName", "") or "Unassigned",
            contact.get("accountDomain", ""),
            contact.get("contactName", ""),
            contact.get("email", ""),
            contact.get("title", ""),
            contact.get("roleBucket", ""),
            contact.get("sourceSystem", ""),
            contact.get("relationshipStatus", ""),
            contact.get("lastTouchAt", ""),
            contact.get("nextStep", ""),
            "Yes" if contact.get("selectedByAm") else "",
            contact.get("notes", ""),
        ]
        for col, value in enumerate(values, 1):
            contacts_sheet.cell(row=row_index, column=col, value=value)
    style_table(contacts_sheet, 3, len(active_contacts) + 3, 1, len(contact_headers))
    contacts_sheet.auto_filter.ref = f"A3:L{len(active_contacts) + 3}"
else:
    contacts_sheet["A4"] = "No active contacts imported yet."
    style_table(contacts_sheet, 3, 4, 1, len(contact_headers))
contacts_sheet.freeze_panes = "A4"
set_widths(contacts_sheet, [30, 26, 28, 34, 34, 22, 18, 20, 20, 36, 12, 44])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)
